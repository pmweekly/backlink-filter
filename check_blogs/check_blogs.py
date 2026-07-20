#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
批量检测 URL 是否为博客网站（含完整评论表单）
输入：Excel 文件，列结构：
  Page address | Source title | Source url | Target url | Anchor |
  External links | Internal links | Nofollow | Sponsored | Ugc |
  Text | Frame | Form | Image | Sitewide | First seen | Last seen |
  New link | Lost link
检测列：Source url（第3列，索引2）
输出：原有列保持不变，在 Lost link 后追加两列：
  - 顶级域名
  - 评论表单检测（博客网站 / 评论网站 / 无评论功能 / 访问失败）
重复顶级域名（Source url 提取）：直接删除该行
"""

import sys
import time
import argparse
import os
import re
import sqlite3
import threading
from dataclasses import asdict, dataclass
from concurrent.futures import Future, ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Callable, Optional, Tuple, List, Set
import csv, json, tempfile
from datetime import datetime
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup


LogCallback = Callable[[str], None]
ProgressCallback = Callable[[float, str], None]
StatsCallback = Callable[[dict[str, Any]], None]


@dataclass
class CheckBlogsResult:
    output_file: str
    output_csv: str
    checkpoint_file: str
    input_file: str
    raw_rows: int
    processed_rows: int
    skipped_duplicate_domains: int
    empty_url_rows: int
    label_counts: dict[str, int]
    network_checked_rows: int = 0
    cache_hit_rows: int = 0
    resumed_rows: int = 0
    google_login_rows: int = 0

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def emit_log(logger: LogCallback | None, message: str) -> None:
    if logger:
        logger(message)


def emit_progress(progress_callback: ProgressCallback | None, percent: float, message: str) -> None:
    if progress_callback:
        progress_callback(max(0.0, min(100.0, percent)), message)


def _iter_rows(input_path: str):
    """流式迭代输入文件的行。首次 yield header（list），随后逐行 yield list(row_values)。"""
    if input_path.lower().endswith('.xls'):
        wb = xlrd.open_workbook(input_path)
        ws = wb.sheet_by_index(0)
        header = [str(ws.cell_value(0, c)) if ws.cell_value(0, c) != '' else '' for c in range(ws.ncols)]
        yield header
        for r in range(1, ws.nrows):
            yield [ws.cell_value(r, c) for c in range(ws.ncols)]
    else:
        wb = openpyxl.load_workbook(input_path, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return
        yield [cell for cell in header]
        for row in rows:
            yield list(row)


import tldextract
import openpyxl
import xlrd
from openpyxl.styles import PatternFill

DOMAIN_EXTRACTOR = tldextract.TLDExtract(suffix_list_urls=(), cache_dir=None)

# ── 请求配置 ────────────────────────────────────────────────────────────────
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}
CONNECT_TIMEOUT = float(os.getenv("CHECK_BLOGS_CONNECT_TIMEOUT_SECONDS", "5"))
READ_TIMEOUT = float(os.getenv("CHECK_BLOGS_READ_TIMEOUT_SECONDS", "20"))
TOTAL_TIMEOUT = float(os.getenv("CHECK_BLOGS_TOTAL_TIMEOUT_SECONDS", "30"))
MAX_RESPONSE_BYTES = int(os.getenv("CHECK_BLOGS_MAX_RESPONSE_MB", "5")) * 1024 * 1024
RETRY_COUNT     = 1    # 超时/失败不重试，快速失败避免浪费时间
DEFAULT_CONCURRENCY = max(1, int(os.getenv("CHECK_BLOGS_CONCURRENCY", "16")))
DOMAIN_INTERVAL_SECONDS = float(os.getenv("CHECK_BLOGS_DOMAIN_INTERVAL_SECONDS", "1.5"))
CHECKPOINT_BATCH_SIZE = max(1, int(os.getenv("CHECK_BLOGS_CHECKPOINT_BATCH_SIZE", "25")))
CACHE_SUCCESS_TTL_SECONDS = int(os.getenv("CHECK_BLOGS_CACHE_SUCCESS_TTL_SECONDS", "604800"))
CACHE_FAILURE_TTL_SECONDS = int(os.getenv("CHECK_BLOGS_CACHE_FAILURE_TTL_SECONDS", "21600"))
CACHE_RULE_VERSION = "2026-07-google-comment-v1"

_thread_local = threading.local()
_executor_lock = threading.Lock()
_shared_executor: ThreadPoolExecutor | None = None
_shared_executor_size = 0
_domain_locks_guard = threading.Lock()
_domain_locks: dict[str, threading.Lock] = {}
_domain_last_started: dict[str, float] = {}


def _get_session() -> requests.Session:
    session = getattr(_thread_local, "session", None)
    if session is None:
        session = requests.Session()
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=DEFAULT_CONCURRENCY,
            pool_maxsize=DEFAULT_CONCURRENCY,
        )
        session.mount("http://", adapter)
        session.mount("https://", adapter)
        _thread_local.session = session
    return session


def get_shared_executor(max_workers: int = DEFAULT_CONCURRENCY) -> ThreadPoolExecutor:
    global _shared_executor, _shared_executor_size
    max_workers = max(1, int(max_workers))
    with _executor_lock:
        if _shared_executor is None:
            _shared_executor = ThreadPoolExecutor(
                max_workers=max_workers,
                thread_name_prefix="check-blogs",
            )
            _shared_executor_size = max_workers
        return _shared_executor


def _get_domain_lock(domain: str) -> threading.Lock:
    with _domain_locks_guard:
        return _domain_locks.setdefault(domain, threading.Lock())


@dataclass
class DetectionResult:
    domain: str
    label: str
    flags: str
    comment_time: str
    final_url: str
    outcome: str
    cache_hit: bool = False


class ProcessingPaused(Exception):
    """Raised after durable progress is flushed for a user-requested pause."""


class DomainResultCache:
    def __init__(self, path: str | None):
        self.path = Path(path) if path else None
        if self.path:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            self._initialize()

    def _connect(self):
        assert self.path is not None
        connection = sqlite3.connect(self.path, timeout=30)
        connection.row_factory = sqlite3.Row
        return connection

    def _initialize(self) -> None:
        with self._connect() as connection:
            connection.execute("PRAGMA journal_mode=WAL")
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS domain_results (
                    domain TEXT PRIMARY KEY,
                    label TEXT NOT NULL,
                    flags TEXT NOT NULL,
                    comment_time TEXT NOT NULL,
                    final_url TEXT NOT NULL,
                    outcome TEXT NOT NULL,
                    checked_at REAL NOT NULL,
                    rule_version TEXT NOT NULL
                )
                """
            )

    def get(self, domain: str) -> DetectionResult | None:
        if not self.path:
            return None
        with self._connect() as connection:
            row = connection.execute(
                "SELECT * FROM domain_results WHERE domain = ? AND rule_version = ?",
                (domain, CACHE_RULE_VERSION),
            ).fetchone()
        if not row:
            return None
        ttl = CACHE_FAILURE_TTL_SECONDS if row["outcome"] in {"timeout", "error"} else CACHE_SUCCESS_TTL_SECONDS
        if time.time() - float(row["checked_at"]) > ttl:
            return None
        return DetectionResult(
            domain=domain,
            label=row["label"],
            flags=row["flags"],
            comment_time=row["comment_time"],
            final_url=row["final_url"],
            outcome=row["outcome"],
            cache_hit=True,
        )

    def put(self, result: DetectionResult) -> None:
        if not self.path:
            return
        with self._connect() as connection:
            connection.execute(
                """
                INSERT INTO domain_results
                    (domain, label, flags, comment_time, final_url, outcome, checked_at, rule_version)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(domain) DO UPDATE SET
                    label=excluded.label, flags=excluded.flags,
                    comment_time=excluded.comment_time, final_url=excluded.final_url,
                    outcome=excluded.outcome, checked_at=excluded.checked_at,
                    rule_version=excluded.rule_version
                """,
                (
                    result.domain, result.label, result.flags, result.comment_time,
                    result.final_url, result.outcome, time.time(), CACHE_RULE_VERSION,
                ),
            )


def load_checkpoint(checkpoint_path: str, out_csv_path: str, logger: LogCallback | None = print) -> Set[str]:
    """合并 checkpoint 和 CSV 中的已处理域名，容忍两次原子写之间中断。"""
    processed = set()
    if os.path.exists(checkpoint_path):
        try:
            with open(checkpoint_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            for d in data.get('processed_domains', []):
                if d:
                    processed.add(str(d).strip().lower())
            emit_log(logger, f"[续跑] 从检查点加载 {len(processed)} 个域名")
        except Exception as e:
            emit_log(logger, f"[WARN] 无法加载 checkpoint {checkpoint_path}: {e}")

    # fallback: 从已有 CSV 重建
    if os.path.exists(out_csv_path):
        try:
            with open(out_csv_path, newline='', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                headers = next(reader, None)
                if headers:
                    headers_l = [h.strip().lower() if h else '' for h in headers]
                    # 尝试定位中文列名 '顶级域名'
                    idx = -1
                    for i, h in enumerate(headers_l):
                        if '顶级域名' in h:
                            idx = i
                            break
                    if idx >= 0:
                        for row in reader:
                            if len(row) > idx and row[idx]:
                                processed.add(row[idx].strip().lower())
                        emit_log(logger, f"[续跑] 合并 CSV 后共 {len(processed)} 个域名")
        except Exception as e:
            emit_log(logger, f"[WARN] 无法从 CSV 重建 checkpoint: {e}")

    # fallback2: 从已有 XLSX 重建（兼容旧结果文件）
    try:
        out_xlsx = os.path.splitext(out_csv_path)[0] + '.xlsx'
        if os.path.exists(out_xlsx):
            wb = openpyxl.load_workbook(out_xlsx, read_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value is not None else '' for c in ws[1]]
            idx = -1
            for i, h in enumerate(headers):
                if '顶级域名' in h:
                    idx = i
                    break
            if idx >= 0:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) > idx and row[idx]:
                        processed.add(str(row[idx]).strip().lower())
                emit_log(logger, f"[续跑] 合并 XLSX 后共 {len(processed)} 个域名")
    except Exception as e:
        emit_log(logger, f"[WARN] 无法从 XLSX 重建 checkpoint: {e}")

    return processed


def get_paths_for_output(input_path: str, output_dir: str | None = None):
    # CLI 默认使用脚本所在目录的 output/ 文件夹；Web 调用可指定任务输出目录。
    if output_dir is None:
        project_root = os.path.dirname(os.path.abspath(__file__))
        output_dir = os.path.join(project_root, "output")
    os.makedirs(output_dir, exist_ok=True)
    base_name = os.path.basename(input_path)
    if base_name.lower().endswith('.xlsx'):
        base = base_name[:-5]
    elif base_name.lower().endswith('.xls'):
        base = base_name[:-4]
    else:
        base = os.path.splitext(base_name)[0]
    out_csv = os.path.join(output_dir, f"{base}_result.csv")
    checkpoint = os.path.join(output_dir, f"{base}_checkpoint.json")
    return out_csv, checkpoint


def resolve_output_paths(
    input_path: str,
    output_path: str | None = None,
    output_csv_path: str | None = None,
    checkpoint_path: str | None = None,
    output_dir: str | None = None,
) -> tuple[str, str, str]:
    if output_path:
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        if output.suffix.lower() == ".csv":
            out_csv = output
            out_xlsx = output.with_suffix(".xlsx")
        else:
            out_xlsx = output if output.suffix else output.with_suffix(".xlsx")
            out_csv = Path(output_csv_path) if output_csv_path else out_xlsx.with_suffix(".csv")
    else:
        out_csv_text, checkpoint_text = get_paths_for_output(input_path, output_dir)
        out_csv = Path(output_csv_path) if output_csv_path else Path(out_csv_text)
        out_xlsx = out_csv.with_suffix(".xlsx")
        checkpoint_path = checkpoint_path or checkpoint_text

    if output_csv_path:
        out_csv = Path(output_csv_path)
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    checkpoint = Path(checkpoint_path) if checkpoint_path else out_csv.with_name(f"{out_csv.stem}_checkpoint.json")
    checkpoint.parent.mkdir(parents=True, exist_ok=True)
    return str(out_csv), str(out_xlsx), str(checkpoint)


def save_checkpoint_atomic(checkpoint_path: str, processed_domains: Set[str]):
    """原子性保存 checkpoint（写入临时文件后替换）。"""
    dirpath = os.path.dirname(checkpoint_path) or '.'
    tmp = None
    try:
        fd, tmp = tempfile.mkstemp(dir=dirpath, prefix='.ckpt', text=True)
        with os.fdopen(fd, 'w', encoding='utf-8') as f:
            json.dump({'processed_domains': list(processed_domains)}, f, ensure_ascii=False)
        os.replace(tmp, checkpoint_path)
    except Exception as e:
        print(f"[WARN] 无法保存 checkpoint {checkpoint_path}: {e}")
        if tmp and os.path.exists(tmp):
            try:
                os.remove(tmp)
            except Exception:
                pass
DELAY_BETWEEN   = 0.0  # 保留 CLI 兼容参数；限速现由同域名调度器控制


# ── 评论表单检测结果颜色映射 ────────────────────────────────────────────────
# 用于生成带颜色标注的 XLSX 输出
LABEL_COLORS = {
    '博客网站':           'C6EFCE',  # 绿色  — 可直接留言
    '博客网站 · 谷歌登录': 'FFEB9C',  # 黄色  — 需谷歌账号
    '博客网站 · 需注册':   'FFEB9C',  # 黄色  — 需注册账号
    '博客网站 · 需登录':   'FFEB9C',  # 黄色  — 需登录
    '评论网站':           'BDD7EE',  # 蓝色  — 评论区不完整
    '评论网站 · 谷歌登录': 'FFEB9C',  # 黄色  — 评论区需 Google 登录
    '无评论功能':          'D9D9D9',  # 灰色  — 无评论区
    '游戏站':             'FCE4D6',  # 橙色  — 游戏类站点
    '访问超时':           'FFCCCC',  # 浅红  — 响应超时
    '访问失败':           'FF9999',  # 红色  — 无法连接
}


# ── 表单检测关键词 ──────────────────────────────────────────────────────────
# 完整博客评论表单需同时满足以下 5 个条件
REQUIRED_SIGNALS = [
    "comment",   # Comment 文本域
    "name",      # Name 输入框
    "email",     # Email 输入框
    "website",   # Website / URL 输入框
    "post comment",  # 提交按钮
]

# 部分评论信号（有评论区但表单不完整）
PARTIAL_SIGNALS = [
    "comment",
    "leave a comment",
    "add a comment",
    "reply",
    "post a comment",
]

# 游戏站关键词（在页面标题/meta/h1/nav 中匹配）
GAME_KEYWORDS = [
    "game", "games", "gaming", "gamer", "gamers",
    "esports", "e-sports", "gameplay", "walkthrough",
    "cheat", "cheats", "cheat code", "cheat codes",
    "mod", "mods", "modding",
    "rpg", "fps", "mmorpg", "moba", "battle royale",
    "online game", "mobile game", "pc game", "video game",
    "xbox", "playstation", "nintendo", "steam",
    "slot", "casino", "poker",
    "download game", "free game", "play game",
]

# 评论需要登录/注册的信号
LOGIN_SIGNALS = [
    "you must be logged in",
    "you must log in",
    "please log in",
    "login to comment",
    "log in to comment",
    "sign in to comment",
    "must be signed in",
]

REGISTER_SIGNALS = [
    "register to comment",
    "sign up to comment",
    "create an account",
    "registration required",
    "sign up for an account",
    "join to comment",
]

GOOGLE_SIGNALS = [
    "sign in with google",
    "signin with google",
    "g_id_onload",
    "accounts.google.com/gsi",
    "googlesignin",
    "google sign-in",
    "google sign in",
    "gsi/client",
]


def get_top_domain(url: str) -> str:
    """返回 URL 的一级注册域名，如 example.com"""
    ext = DOMAIN_EXTRACTOR(url)
    if ext.domain and ext.suffix:
        return f"{ext.domain}.{ext.suffix}"
    # 回退：直接用 hostname
    parsed = urlparse(url)
    return parsed.netloc.lower()


def fetch_page(url: str, logger: LogCallback | None = print) -> Tuple[Optional[str], str, str]:
    """
    获取页面 HTML。
    返回: (html, 最终URL, 错误类型)
      错误类型: ""        — 成功
               "timeout" — 请求超时（超过 REQUEST_TIMEOUT）
               "error"   — 其他访问失败
    """
    for attempt in range(RETRY_COUNT):
        try:
            started = time.monotonic()
            resp = _get_session().get(
                url,
                headers=HEADERS,
                timeout=(CONNECT_TIMEOUT, READ_TIMEOUT),
                allow_redirects=True,
                stream=True,
            )
            resp.raise_for_status()
            body = bytearray()
            for chunk in resp.iter_content(chunk_size=64 * 1024):
                if time.monotonic() - started > TOTAL_TIMEOUT:
                    raise requests.Timeout(f"整体请求超过 {TOTAL_TIMEOUT}s")
                if not chunk:
                    continue
                body.extend(chunk)
                if len(body) > MAX_RESPONSE_BYTES:
                    emit_log(logger, f"  [LIMIT] {url} 响应体超过 {MAX_RESPONSE_BYTES // 1024 // 1024}MB，已截断")
                    break
            encoding = resp.encoding or "utf-8"
            return bytes(body).decode(encoding, errors="replace"), resp.url, ""
        except requests.Timeout:
            emit_log(logger, f"  [TIMEOUT] {url} 连接超时 {CONNECT_TIMEOUT}s / 读取超时 {READ_TIMEOUT}s")
            return None, url, "timeout"
        except requests.RequestException as e:
            if attempt < RETRY_COUNT - 1:
                time.sleep(2)
            else:
                emit_log(logger, f"  [ERROR] {url} -> {e}")
                return None, url, "error"
    return None, url, "error"


def detect_url(
    url: str,
    domain: str,
    cache: DomainResultCache,
    logger: LogCallback | None = print,
    domain_interval: float = DOMAIN_INTERVAL_SECONDS,
) -> DetectionResult:
    cached = cache.get(domain)
    if cached:
        return cached

    domain_lock = _get_domain_lock(domain)
    with domain_lock:
        cached = cache.get(domain)
        if cached:
            return cached

        with _domain_locks_guard:
            wait_for = max(0.0, domain_interval - (time.monotonic() - _domain_last_started.get(domain, 0.0)))
        if wait_for:
            time.sleep(wait_for)
        with _domain_locks_guard:
            _domain_last_started[domain] = time.monotonic()

        html, final_url, err_type = fetch_page(url, logger=logger)
        flags: list[str] = []
        comment_time = ""
        if html is None:
            label = '访问超时' if err_type == "timeout" else '访问失败'
            outcome = err_type or "error"
        elif detect_game_site(html, url):
            label = '游戏站'
            flags.append('游戏网站')
            outcome = "success"
        else:
            form_result = detect_comment_form(html)
            login_type = detect_comment_login(html) if form_result in {"blog", "comment"} else "open"
            if form_result == "blog":
                if login_type == "google":
                    label = '博客网站 · 谷歌登录'
                    flags.append('需登录（谷歌登录）')
                elif login_type == "register":
                    label = '博客网站 · 需注册'
                    flags.append('需登录')
                elif login_type == "login":
                    label = '博客网站 · 需登录'
                    flags.append('需登录')
                else:
                    label = '博客网站'
                comment_time = get_latest_comment_time(html) or ""
            elif form_result == "comment":
                if login_type == "google":
                    label = '评论网站 · 谷歌登录'
                    flags.append('需登录（谷歌登录）')
                else:
                    label = '评论网站'
            else:
                label = '无评论功能'
            outcome = "success"

        result = DetectionResult(
            domain=domain,
            label=label,
            flags='；'.join(flags),
            comment_time=comment_time,
            final_url=final_url,
            outcome=outcome,
        )
        cache.put(result)
        return result


def _text_contains_any(text: str, keywords) -> bool:
    text_lower = text.lower()
    return any(kw in text_lower for kw in keywords)


def _collect_form_text(form) -> str:
    """提取表单内所有文本、label、placeholder、name、id、value 等"""
    parts = []
    for tag in form.find_all(True):
        for attr in ("placeholder", "name", "id", "value", "aria-label", "title"):
            val = tag.get(attr, "")
            if val:
                parts.append(val)
        text = tag.get_text(" ", strip=True)
        if text:
            parts.append(text)
    return " ".join(parts)


def _check_submit_button(form) -> bool:
    """检查表单内是否存在 'Post Comment'（或变体）提交按钮"""
    for btn in form.find_all(["button", "input"]):
        btn_text = (
            btn.get_text(" ", strip=True)
            + " "
            + btn.get("value", "")
            + " "
            + btn.get("id", "")
            + " "
            + btn.get("name", "")
        ).lower()
        if "post comment" in btn_text or "submit" in btn_text or "post" in btn_text:
            return True
    return False


def detect_game_site(html: str, url: str) -> bool:
    """
    检测页面是否为游戏站。
    命中逻辑：
      - URL 域名含 game/games/gaming → 直接判定
      - 否则在 <title>, <meta keywords/description>, <h1>, <nav> 中匹配
        GAME_KEYWORDS，命中 2 处及以上则判定
    """
    # 先检查域名
    domain = get_top_domain(url).lower()
    if any(kw in domain for kw in ("game", "games", "gaming")):
        return True

    soup = BeautifulSoup(html, "html.parser")

    def _score(text: str) -> int:
        text_lower = text.lower()
        return sum(1 for kw in GAME_KEYWORDS if kw in text_lower)

    hit_count = 0

    title_tag = soup.find("title")
    if title_tag:
        hit_count += min(_score(title_tag.get_text()), 1)

    for meta in soup.find_all("meta"):
        name = (meta.get("name") or "").lower()
        if name in ("keywords", "description"):
            content = meta.get("content") or ""
            hit_count += min(_score(content), 1)

    for h1 in soup.find_all("h1"):
        hit_count += min(_score(h1.get_text()), 1)

    for nav in soup.find_all("nav"):
        hit_count += min(_score(nav.get_text()), 1)

    return hit_count >= 2


def detect_comment_login(html: str) -> str:
    """
    检测博客评论区是否有登录/注册门槛。
    返回值（优先级从高到低）：
      "google"   — 需要 Google 账号登录
      "register" — 需要注册账号
      "login"    — 需要登录（第三方插件或通用登录墙）
      "open"     — 无门槛，可直接评论
    """
    html_lower = html.lower()

    # Google 登录检测（直接扫描原始 HTML，包含 JS 片段）
    if any(sig in html_lower for sig in GOOGLE_SIGNALS):
        return "google"

    # Disqus / Facebook 评论插件 → 需登录
    if "disqus.com/embed" in html_lower or "fb-comments" in html_lower or "facebook.com/plugins/comments" in html_lower:
        return "login"

    soup = BeautifulSoup(html, "html.parser")
    page_text = soup.get_text(" ", strip=True)

    if _text_contains_any(page_text, REGISTER_SIGNALS):
        return "register"

    if _text_contains_any(page_text, LOGIN_SIGNALS):
        return "login"

    return "open"


def _parse_date_str(s: str) -> Optional[datetime]:
    """尝试将各种格式的日期字符串解析为 datetime，解析失败返回 None。"""
    s = s.strip()
    if not s:
        return None
    # ISO 8601 / RFC 3339（去掉时区后缀后再解析）
    try:
        clean = re.sub(r'[+-]\d{2}:\d{2}$', '', s).rstrip('Z')
        return datetime.fromisoformat(clean)
    except Exception:
        pass
    # 常见博客日期格式
    for fmt in (
        '%B %d, %Y at %I:%M %p',  # April 27, 2026 at 10:30 am
        '%B %d, %Y at %H:%M',     # April 27, 2026 at 10:30
        '%B %d, %Y',              # April 27, 2026
        '%b %d, %Y',              # Apr 27, 2026
        '%Y-%m-%dT%H:%M:%S',
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M',
        '%Y-%m-%d',
        '%d/%m/%Y',
        '%m/%d/%Y',
        '%Y/%m/%d',
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def get_latest_comment_time(html: str) -> Optional[str]:
    """
    从博客页面提取最近一条评论的发布时间。
    策略（优先级从高到低）：
      1. JSON-LD 结构化数据中 @type=Comment 的 dateCreated/datePublished
      2. 评论容器内 <time datetime="..."> 标签
      3. 常见博客评论日期类名选择器
    返回格式化为 'YYYY-MM-DD HH:MM' 的字符串，或 None。
    """
    soup = BeautifulSoup(html, "html.parser")
    dates: List[str] = []

    # 策略1: JSON-LD
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or '')
            items = data if isinstance(data, list) else [data]
            for item in items:
                if not isinstance(item, dict):
                    continue
                if item.get("@type") == "Comment":
                    d = item.get("dateCreated") or item.get("datePublished")
                    if d:
                        dates.append(str(d))
                # Article/BlogPosting 内嵌的 comment 列表
                for c in (item.get("comment") or []):
                    if isinstance(c, dict):
                        d = c.get("dateCreated") or c.get("datePublished")
                        if d:
                            dates.append(str(d))
        except Exception:
            pass

    # 策略2: 评论容器内的 <time> 标签
    for container in soup.find_all(class_=re.compile(r'\bcomments?\b', re.I)):
        for time_tag in container.find_all("time"):
            d = time_tag.get("datetime") or time_tag.get_text(strip=True)
            if d:
                dates.append(d)

    # 策略3: 常见博客评论日期选择器
    for selector in (
        '.comment-metadata time',
        '.comment-meta time',
        '.comment-date',
        '.commentmetadata',
        '.comment-time',
        'footer.comment-meta time',
        'article.comment time',
    ):
        for elem in soup.select(selector):
            d = elem.get("datetime") or elem.get_text(strip=True)
            if d:
                dates.append(d)

    if not dates:
        return None

    # 解析后取最新时间
    parsed: List[Tuple[datetime, str]] = []
    for d in dates:
        dt = _parse_date_str(d)
        if dt:
            parsed.append((dt, d))

    if parsed:
        parsed.sort(key=lambda x: x[0], reverse=True)
        return parsed[0][0].strftime("%Y-%m-%d %H:%M")

    # 无法解析则返回第一个原始字符串
    return dates[0]


def detect_comment_form(html: str) -> str:
    """
    返回值：
      "blog"    — 存在完整评论表单
      "comment" — 存在部分评论区但不完整
      "none"    — 未检测到评论功能
    """
    soup = BeautifulSoup(html, "html.parser")

    for form in soup.find_all("form"):
        form_text = _collect_form_text(form)

        has_comment = _text_contains_any(form_text, ["comment"])
        has_name    = _text_contains_any(form_text, ["name"])
        has_email   = _text_contains_any(form_text, ["email"])
        has_website = _text_contains_any(form_text, ["website", "url"])
        has_btn     = _check_submit_button(form)

        if has_comment and has_name and has_email and has_website and has_btn:
            return "blog"

        if has_comment or _text_contains_any(form_text, ["leave a comment", "add a comment"]):
            return "comment"

    page_text = soup.get_text(" ", strip=True)
    if _text_contains_any(page_text, PARTIAL_SIGNALS):
        return "comment"

    return "none"


def write_colored_xlsx(csv_path: str, xlsx_path: str, logger: LogCallback | None = print):
    """
    将 CSV 结果转换为带颜色标注的 XLSX 文件。
    根据「评论表单检测」列的值对整行着色，以便快速区分不同类型。
    颜色方案：
      绿色  — 博客网站（可直接留言）
      黄色  — 博客网站（需登录/注册/谷歌账号）
      蓝色  — 评论网站（不完整评论区）
      灰色  — 无评论功能
      橙色  — 游戏站
      浅红  — 访问超时
      红色  — 访问失败
    """
    if not os.path.exists(csv_path):
        emit_log(logger, f"[WARN] CSV 文件不存在，跳过生成 XLSX: {csv_path}")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    label_col_idx = -1  # 「评论表单检测」列的 1-based 列号

    with open(csv_path, newline='', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, 1):
            ws.append(row)
            if row_idx == 1:
                # 定位「评论表单检测」列
                for ci, h in enumerate(row):
                    if '评论表单检测' in (h or ''):
                        label_col_idx = ci + 1  # openpyxl 列号从 1 开始
                        break
            elif label_col_idx > 0 and len(row) >= label_col_idx:
                label_val = row[label_col_idx - 1]
                color = LABEL_COLORS.get(label_val)
                if color:
                    fill = PatternFill(fill_type="solid", fgColor=color)
                    for col in range(1, len(row) + 1):
                        ws.cell(row=row_idx, column=col).fill = fill

    # 冻结首行、自动列宽（近似）
    ws.freeze_panes = "A2"
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                val_len = len(str(cell.value)) if cell.value is not None else 0
                if val_len > max_len:
                    max_len = val_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    wb.save(xlsx_path)
    emit_log(logger, f"[XLSX] 带颜色标注的 Excel 已生成: {xlsx_path}")


def _read_all_rows(input_path: str):
    """读取 .xls 或 .xlsx，返回 (header_list, data_rows_list)，均为纯值列表"""
    if input_path.lower().endswith(".xls"):
        wb = xlrd.open_workbook(input_path)
        ws = wb.sheet_by_index(0)
        header = [str(ws.cell_value(0, c)) if ws.cell_value(0, c) != '' else '' for c in range(ws.ncols)]
        data_rows = []
        for r in range(1, ws.nrows):
            data_rows.append([ws.cell_value(r, c) for c in range(ws.ncols)])
        return header, data_rows
    else:
        wb = openpyxl.load_workbook(input_path)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        data_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data_rows.append(list(row))
        return header, data_rows


def count_data_rows(input_path: str) -> int:
    if input_path.lower().endswith(".xls"):
        wb = xlrd.open_workbook(input_path)
        ws = wb.sheet_by_index(0)
        return max(ws.nrows - 1, 0)
    wb = openpyxl.load_workbook(input_path, read_only=True)
    ws = wb.active
    return max((ws.max_row or 1) - 1, 0)


def sort_csv_by_input_order(input_path: str, csv_path: str) -> None:
    """并发期间允许完成结果即时落盘，最终导出前按原始输入顺序重排。"""
    order: dict[str, int] = {}
    input_rows = _iter_rows(input_path)
    try:
        input_header = [str(value or "").strip().lower() for value in next(input_rows)]
    except StopIteration:
        return
    source_idx = next((i for i, value in enumerate(input_header) if value == "source url"), 2)
    for index, row in enumerate(input_rows):
        raw_url = row[source_idx] if source_idx < len(row) else ""
        url = str(raw_url).strip() if raw_url else ""
        if not url:
            continue
        domain = get_top_domain(url).strip().lower() or urlparse(url).netloc.lower()
        order.setdefault(domain, index)

    with open(csv_path, newline="", encoding="utf-8-sig") as source:
        reader = csv.reader(source)
        header = next(reader, [])
        rows = list(reader)
    domain_idx = next((i for i, value in enumerate(header) if "顶级域名" in (value or "")), -1)
    if domain_idx < 0:
        return
    rows.sort(
        key=lambda row: order.get(
            row[domain_idx].strip().lower() if len(row) > domain_idx else "",
            len(order) + 1,
        )
    )
    directory = os.path.dirname(csv_path) or "."
    fd, temp_path = tempfile.mkstemp(dir=directory, prefix=".sorted", text=True)
    try:
        with os.fdopen(fd, "w", newline="", encoding="utf-8-sig") as destination:
            writer = csv.writer(destination)
            writer.writerow(header)
            writer.writerows(rows)
            destination.flush()
            os.fsync(destination.fileno())
        os.replace(temp_path, csv_path)
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


def process_excel(
    input_path: str,
    output_path: str | None = None,
    *,
    output_csv_path: str | None = None,
    checkpoint_path: str | None = None,
    output_dir: str | None = None,
    resume: bool = True,
    logger: LogCallback | None = print,
    progress_callback: ProgressCallback | None = None,
    delay_between: float = DELAY_BETWEEN,
    max_workers: int = DEFAULT_CONCURRENCY,
    checkpoint_batch_size: int = CHECKPOINT_BATCH_SIZE,
    cache_path: str | None = None,
    domain_interval: float = DOMAIN_INTERVAL_SECONDS,
    stats_callback: StatsCallback | None = None,
    should_pause: Callable[[], bool] | None = None,
) -> CheckBlogsResult:
    """并发检测 URL，按输入顺序批量写入 CSV/XLSX，并支持断点续跑。"""
    out_csv, out_xlsx, checkpoint = resolve_output_paths(
        input_path,
        output_path=output_path,
        output_csv_path=output_csv_path,
        checkpoint_path=checkpoint_path,
        output_dir=output_dir,
    )
    total_rows = count_data_rows(input_path)
    emit_progress(progress_callback, 2, "正在读取 check_blogs 输入")

    if not resume:
        for path in (out_csv, out_xlsx, checkpoint):
            try:
                if os.path.exists(path):
                    os.remove(path)
            except Exception as exc:
                emit_log(logger, f"[WARN] 无法删除旧输出 {path}: {exc}")

    rows = _iter_rows(input_path)
    try:
        header = next(rows)
    except StopIteration:
        emit_log(logger, "[ERROR] 输入文件为空，退出")
        emit_progress(progress_callback, 100, "check_blogs 输入为空")
        return CheckBlogsResult(
            output_file="",
            output_csv=out_csv,
            checkpoint_file=checkpoint,
            input_file=input_path,
            raw_rows=0,
            processed_rows=0,
            skipped_duplicate_domains=0,
            empty_url_rows=0,
            label_counts={},
        )

    header_row = [str(h) if h is not None else '' for h in header]

    try:
        source_url_col_idx = next(
            i for i, h in enumerate(header_row) if h and str(h).strip().lower() == "source url"
        )
    except StopIteration:
        source_url_col_idx = 2
        emit_log(logger, "[WARNING] 未找到 'Source url' 表头，默认使用第3列")

    csv_exists = resume and os.path.exists(out_csv)
    mode = 'a' if csv_exists else 'w'
    processed_domains = load_checkpoint(checkpoint, out_csv, logger=logger) if resume else set()
    seen_in_run: Set[str] = set()
    resumed_rows = 0
    processed_now = 0
    skipped_duplicate_domains = 0
    empty_url_rows = 0
    label_counts: dict[str, int] = {}
    network_checked_rows = 0
    cache_hit_rows = 0
    google_login_rows = 0
    resumed_empty_rows = 0

    if csv_exists:
        try:
            with open(out_csv, newline='', encoding='utf-8-sig') as existing_file:
                existing_reader = csv.reader(existing_file)
                existing_header = next(existing_reader, [])
                label_idx = next((i for i, value in enumerate(existing_header) if '评论表单检测' in (value or '')), -1)
                domain_idx = next((i for i, value in enumerate(existing_header) if '顶级域名' in (value or '')), -1)
                for existing_row in existing_reader:
                    resumed_rows += 1
                    if label_idx >= 0 and len(existing_row) > label_idx:
                        existing_label = existing_row[label_idx]
                        label_counts[existing_label] = label_counts.get(existing_label, 0) + 1
                        if '谷歌登录' in existing_label:
                            google_login_rows += 1
                    if domain_idx >= 0 and (len(existing_row) <= domain_idx or not existing_row[domain_idx]):
                        resumed_empty_rows += 1
        except Exception as exc:
            emit_log(logger, f"[WARN] 无法重建续跑统计: {exc}")

    cache = DomainResultCache(cache_path or os.getenv("CHECK_BLOGS_CACHE_PATH"))
    executor = get_shared_executor(max_workers)
    work_items: list[tuple[int, list[Any], str, Future[DetectionResult] | None]] = []

    for row_idx, row_values in enumerate(rows, 1):
        row_values = list(row_values) if row_values is not None else []
        raw_url = row_values[source_url_col_idx] if source_url_col_idx < len(row_values) else ""
        url = str(raw_url).strip() if raw_url else ""
        if not url:
            if resumed_empty_rows > 0:
                resumed_empty_rows -= 1
                continue
            work_items.append((row_idx, row_values, "", None))
            continue
        domain = get_top_domain(url).strip().lower() or urlparse(url).netloc.lower()
        if domain in processed_domains or domain in seen_in_run:
            skipped_duplicate_domains += 1
            continue
        seen_in_run.add(domain)
        future = executor.submit(detect_url, url, domain, cache, logger, domain_interval)
        work_items.append((row_idx, row_values, domain, future))

    pending_batch = 0
    processed_now = resumed_rows
    with open(out_csv, mode, newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.writer(csvfile)
        if not csv_exists:
            out_header = list(header_row)
            out_header.extend(["顶级域名", "评论表单检测", "特殊标注", "最新评论时间"])
            writer.writerow(out_header)
            csvfile.flush()
            os.fsync(csvfile.fileno())

        def flush_batch() -> None:
            nonlocal pending_batch
            if pending_batch <= 0:
                return
            csvfile.flush()
            os.fsync(csvfile.fileno())
            save_checkpoint_atomic(checkpoint, processed_domains)
            if stats_callback:
                stats_callback({
                    "raw_rows": total_rows,
                    "processed_rows": processed_now,
                    "network_checked_rows": network_checked_rows,
                    "cache_hit_rows": cache_hit_rows,
                    "resumed_rows": resumed_rows,
                    "google_login_rows": google_login_rows,
                    "label_counts": dict(label_counts),
                })
            emit_progress(
                progress_callback,
                5 + (processed_now / max(total_rows, 1)) * 88,
                f"已完成 {processed_now}/{total_rows or '?'} 行（网络 {network_checked_rows}，缓存 {cache_hit_rows}）",
            )
            pending_batch = 0

        try:
            future_items: dict[Future[DetectionResult], tuple[int, list[Any], str]] = {}
            for row_idx, row_values, domain, future in work_items:
                if future is None:
                    out_row = [str(v) if v is not None else '' for v in row_values]
                    out_row.extend(['', '', '', ''])
                    writer.writerow(out_row)
                    processed_now += 1
                    empty_url_rows += 1
                    pending_batch += 1
                    if pending_batch >= checkpoint_batch_size:
                        flush_batch()
                else:
                    future_items[future] = (row_idx, row_values, domain)

            for future in as_completed(future_items):
                if should_pause and should_pause():
                    for pending in future_items:
                        pending.cancel()
                    flush_batch()
                    raise ProcessingPaused("任务已暂停，已保存当前进度。")
                row_idx, row_values, domain = future_items[future]
                result = future.result()
                if result.cache_hit:
                    cache_hit_rows += 1
                else:
                    network_checked_rows += 1
                out_row = [str(v) if v is not None else '' for v in row_values]
                out_row.extend([domain, result.label, result.flags, result.comment_time])
                writer.writerow(out_row)
                processed_domains.add(domain)
                processed_now += 1
                label_counts[result.label] = label_counts.get(result.label, 0) + 1
                if '谷歌登录' in result.label:
                    google_login_rows += 1
                pending_batch += 1
                if pending_batch >= checkpoint_batch_size:
                    flush_batch()
            flush_batch()

        except KeyboardInterrupt:
            flush_batch()
            emit_log(logger, '\n[中断] 已保存当前进度，退出')
            write_colored_xlsx(out_csv, out_xlsx, logger=logger)
            return CheckBlogsResult(
                output_file=out_xlsx,
                output_csv=out_csv,
                checkpoint_file=checkpoint,
                input_file=input_path,
                raw_rows=total_rows,
                processed_rows=processed_now,
                skipped_duplicate_domains=skipped_duplicate_domains,
                empty_url_rows=empty_url_rows,
                label_counts=label_counts,
                network_checked_rows=network_checked_rows,
                cache_hit_rows=cache_hit_rows,
                resumed_rows=resumed_rows,
                google_login_rows=google_login_rows,
            )

    emit_progress(progress_callback, 94, "正在按原始输入顺序整理结果")
    sort_csv_by_input_order(input_path, out_csv)
    emit_progress(progress_callback, 95, "正在生成带颜色标注的 Excel")
    write_colored_xlsx(out_csv, out_xlsx, logger=logger)
    emit_log(logger, f"\n完成！CSV: {out_csv}")
    emit_log(logger, f"Excel（含颜色）: {out_xlsx}")
    emit_progress(progress_callback, 100, "check_blogs 检测完成")
    return CheckBlogsResult(
        output_file=out_xlsx,
        output_csv=out_csv,
        checkpoint_file=checkpoint,
        input_file=input_path,
        raw_rows=total_rows,
        processed_rows=processed_now,
        skipped_duplicate_domains=skipped_duplicate_domains,
        empty_url_rows=empty_url_rows,
        label_counts=label_counts,
        network_checked_rows=network_checked_rows,
        cache_hit_rows=cache_hit_rows,
        resumed_rows=resumed_rows,
        google_login_rows=google_login_rows,
    )


def main():
    parser = argparse.ArgumentParser(
        description="批量检测博客评论表单并输出 CSV（支持断点续跑）"
    )
    parser.add_argument("input", help="输入 Excel 文件路径（.xlsx 或 .xls）")
    args = parser.parse_args()

    input_path = args.input
    out_csv, checkpoint = get_paths_for_output(input_path)
    print(f"输出目录: {os.path.dirname(out_csv)}")

    process_excel(input_path, out_csv, checkpoint_path=checkpoint)


if __name__ == "__main__":
    main()
